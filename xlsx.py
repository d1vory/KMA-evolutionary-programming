"""
    @author: Kyrylo Kundik
    @date: 22-02-2022

    XLSX class for easily generating reports into .xlsx files.
    It can insert rows, columns of data, titles, and apply some styles.
"""
import datetime
import logging

import openpyxl
from openpyxl.styles import Border, Font, NamedStyle, PatternFill, Side


def gen_style(
        name: str,
        bg_color: str = "FFFFFF", color: str = "000000",
        font_size: int = 12, font_style: str = "normal"
) -> NamedStyle:
    """
    Generating new named style for .xlsx file.
    :param name: name of the style
    :param bg_color: background color for cells
    :param color: font color for cells
    :param font_size: size of fonts for cells
    :param font_style: style of fonts for cells
    :return: Named style object that will be registered in the workbook.
    """
    style = NamedStyle(name)
    font = Font(size=font_size, color=color)
    fill = PatternFill("solid", fgColor=bg_color)
    side = Side(style="thin", color="000000")

    if font_style == "bold":
        font.bold = True
    elif font_style == "italic":
        font.italic = True

    style.font = font
    style.fill = fill
    style.border = Border(left=side, right=side, top=side, bottom=side)

    return style


class XLSX:
    def __init__(
            self,
            filename: str = "workbook.xlsx",
            sheet_name: str = "Sheet 1",
            current_row: int = 1,
            current_col: int = 1
    ):
        self._filename = filename
        self._current_row = current_row
        self._current_col = current_col
        self._current_sheet = sheet_name

        self._sheets = {}
        self._workbook = openpyxl.Workbook()

        if "Sheet" in self._workbook.sheetnames:
            del self._workbook["Sheet"]

        self._current_worksheet = None

        self._font_size = 12

        self._logger = logging.getLogger("XLSX_REPORTER")

        self.sheet(sheet_name, current_row, current_col)

        for style in [
            gen_style("normal", bg_color="FFFFFF", color="000000", font_size=self._font_size, font_style="normal"),
            gen_style("bold", bg_color="FFFFFF", color="000000", font_size=self._font_size, font_style="bold"),
            gen_style("bold_bg", bg_color="EDEDED", color="000000", font_size=self._font_size, font_style="bold"),
            gen_style("normal_bg", bg_color="EDEDED", color="000000", font_size=self._font_size, font_style="normal"),
            gen_style("italic_bg", bg_color="EDEDED", color="000000", font_size=self._font_size, font_style="italic"),
            gen_style("italic", bg_color="FFFFFF", color="000000", font_size=self._font_size, font_style="italic")
        ]:
            self._workbook.add_named_style(style)

    def _log(self, mode: str, msg: str, variables: dict = None):
        if not variables:
            variables = {}
        variables["current_timestamp"] = datetime.datetime.now().isoformat()

        variables_repr = " ".join([f"{var}={variables[var]}" for var in variables])

        msg = f" MESSAGE: {msg} | VARIABLES: {variables_repr}"

        if mode == "info":
            self._logger.info(msg)
        elif mode == "warn":
            self._logger.warning(msg)
        elif mode == "debug":
            self._logger.debug(msg)
        elif mode == "error":
            self._logger.error(msg)

    def save(self, filename=None):
        """
        Save .xlsx file with specifying name or will be used
        filename from constructor or default.
        :param filename: Overriding constructor parameter
        :return: None
        """
        filename = filename or self._filename

        if not filename:
            self._log("warning", "Filename for saving was not provided, using default - `workbook.xlsx`")
            filename = "workbook.xlsx"

        self._workbook.save(filename)
        self._log("info", "Workbook successfully saved.", {"filename": filename})

    def sheet(
            self, sheet_name: str, current_row: int = 1, current_col: int = 1
    ):
        """
        Switch to another sheet and save current positions
        :param sheet_name: New sheet to switch
        :param current_row: Start row on a new sheet
        :param current_col: Start column on a new sheet
        :return: None
        """
        self._sheets[self._current_sheet] = (self._current_row, self._current_col)

        self._current_sheet = sheet_name
        self._current_row, self._current_col = (current_row, current_col)

        if sheet_name in self._workbook.sheetnames:
            self._current_worksheet = self._workbook[sheet_name]
        else:
            self._current_worksheet = self._workbook.create_sheet(sheet_name)

        self._log(
            "info", "switched to a new sheet",
            {"sheet_name": sheet_name, "current_row": current_row, "current_col": current_col}
        )

    def skip(self, row: int = 0, col: int = 0):
        self._current_row += row
        self._current_col += col

    def set_pos(self, row: int = None, col: int = None):
        self._current_row = row if row is not None else self._current_row
        self._current_col = col if col is not None else self._current_col

    def text(
            self, title, row=None, col=None,
            override=True, increase_row=True, increase_col=False, style="normal"
    ):
        increase_row = increase_row if row is None else False
        increase_col = increase_col if col is None else False

        row = row or self._current_row

        col = col or self._current_col

        required_cells = len(title) // self._font_size

        end_col = col + required_cells

        for c in range(col, end_col, 1):
            # applying style
            self.cell(
                "", row=row, col=c,
                override=override, increase_row=False, increase_col=False, style=style
            )

        self._current_worksheet.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=end_col
        )

        self.cell(
            title, row=row, col=col,
            override=override, increase_row=False, increase_col=False, style=style
        )

        if increase_row:
            self._current_row += 1
        if increase_col:
            self._current_col += required_cells

    def cell(
            self, value, row=None, col=None,
            override=True, increase_row=True, increase_col=True,
            style="normal"
    ):
        increase_col = increase_col if col is None else False
        increase_row = increase_row if row is None else False

        row = row or self._current_row
        col = col or self._current_col

        cell = self._current_worksheet.cell(row=row, column=col)

        vars_to_log = {"row": row, "col": col, "prev_value": cell.value, "new_value": value}

        if cell.value and not override:
            self._log("warn", "cell was not overwritten", vars_to_log)
        else:
            cell.value = value
            cell.style = style

        if increase_col:
            self._current_col += 1
        if increase_row:
            self._current_row += 1

    def row(
            self, arr, col=None, row=None,
            override=True, increase_row=True, increase_col=False, style="normal"
    ):
        increase_col = increase_col if col is None else False
        increase_row = increase_row if row is None else False

        col = col or self._current_col

        for value in arr:
            self.cell(
                value, row=row, col=col,
                override=override, increase_row=False, increase_col=False, style=style
            )
            col += 1

        if increase_row:
            self._current_row += 1
        if increase_col:
            self._current_col += len(arr)

    def col(
            self, arr, col=None, row=None,
            override=True, increase_col=True, increase_row=False, style="normal"
    ):
        increase_col = increase_col if col is None else False
        increase_row = increase_row if row is None else False

        row = row or self._current_row

        for value in arr:
            self.cell(
                value, row=row, col=col,
                override=override, increase_row=False, increase_col=False, style=style
            )
            row += 1

        if increase_col:
            self._current_col += 1
        if increase_row:
            self._current_row += len(arr)


if __name__ == "__main__":
    reporter = XLSX()
    reporter.row([1, 2, 3, 4, 5, 6, 7, 8, 9, 10], style="bold")
    reporter.row(["123", "234", "345"])
    reporter.skip(5, 2)
    reporter.set_pos(col=1)
    reporter.text("a very long text that I want its to be a great article before my shit that I want to insert below")
    reporter.col(["321", "333", "222"], style="bold_bg")
    reporter.save()
